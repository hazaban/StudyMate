"""Export routes for cards and mistakes (Excel/PDF/CSV)."""

from uuid import UUID
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from jose import jwt

from database import get_db, FlashCard, Mistake, StudyPlan
from config import SECRET_KEY, ALGORITHM

router = APIRouter(prefix="/api/export", tags=["export"])


def _get_user_id(authorization: str = Header(None)) -> UUID:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="未登录")
    try:
        payload = jwt.decode(authorization[7:], SECRET_KEY, algorithms=[ALGORITHM])
        return UUID(payload["sub"])
    except Exception:
        raise HTTPException(status_code=401, detail="登录已过期")


def _escape_csv_field(val):
    if val is None:
        return ""
    s = str(val).replace('"', '""')
    if "," in s or "\n" in s or '"' in s:
        return f'"{s}"'
    return s


def _generate_csv(headers: list, rows: list[list]) -> str:
    lines = [",".join(_escape_csv_field(h) for h in headers)]
    for row in rows:
        lines.append(",".join(_escape_csv_field(c) for c in row))
    return "\n".join(lines)


def _generate_excel(headers: list, rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    """Generate Excel file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F7D4F", end_color="2F7D4F", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_pdf(headers: list, rows: list[list], title: str = "Export") -> bytes:
    """Generate PDF file using reportlab with CJK font support."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # Register CJK font for Chinese text support
    try:
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        cjk_font = 'STSong-Light'
    except Exception:
        cjk_font = 'Helvetica'

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=10, fontName=cjk_font)
    elements = [Paragraph(title, title_style), Spacer(1, 5*mm)]

    # Wrap text in paragraphs for table cells with CJK font
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10, fontName=cjk_font)
    header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=9, leading=11, textColor=colors.white, fontName=cjk_font)

    pdf_data = [[Paragraph(str(h), header_style) for h in headers]]
    for row in rows:
        pdf_data.append([Paragraph(str(c)[:200], cell_style) for c in row])

    col_widths = [min(160*mm / len(headers), 60*mm)] * len(headers)
    table = Table(pdf_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F7D4F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), cjk_font),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbf6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#edf7ee')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 1), (-1, -1), cjk_font),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def _query_cards(db, plan_id, user_id, subject=None, tag=None, mastery_level=None):
    plan = db.query(StudyPlan).filter(StudyPlan.id == plan_id, StudyPlan.user_id == user_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")

    query = db.query(FlashCard).filter(FlashCard.plan_id == plan_id)
    if subject:
        query = query.filter(FlashCard.subject == subject)
    if mastery_level:
        query = query.filter(FlashCard.mastery_level == mastery_level)
    cards = query.order_by(FlashCard.created_at.desc()).all()
    if tag:
        cards = [c for c in cards if c.tags and tag in (c.tags or [])]
    return cards


def _query_mistakes(db, plan_id, user_id, subject=None, tag=None, difficulty=None, mastered=None, min_errors=None):
    plan = db.query(StudyPlan).filter(StudyPlan.id == plan_id, StudyPlan.user_id == user_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")

    query = db.query(Mistake).filter(Mistake.plan_id == plan_id)
    if subject:
        query = query.filter(Mistake.subject == subject)
    if difficulty:
        query = query.filter(Mistake.difficulty == difficulty)
    if mastered is not None:
        query = query.filter(Mistake.mastered == mastered)
    mistakes = query.order_by(Mistake.created_at.desc()).all()
    if tag:
        mistakes = [m for m in mistakes if m.tags and tag in (m.tags or [])]
    if min_errors:
        mistakes = [m for m in mistakes if m.error_count >= int(min_errors)]
    return mistakes


def _card_rows(cards, questions_only=False):
    if questions_only:
        headers = ["科目", "问题", "掌握程度", "标签"]
        mastery_map = {"unmastered": "未掌握", "familiar": "较熟悉", "mastered": "已掌握"}
        rows = []
        for c in cards:
            rows.append([
                c.subject, c.question,
                mastery_map.get(c.mastery_level, c.mastery_level),
                ", ".join(c.tags or []),
            ])
    else:
        headers = ["科目", "问题", "答案", "掌握程度", "复习次数", "标签", "下次复习日期", "创建日期"]
        mastery_map = {"unmastered": "未掌握", "familiar": "较熟悉", "mastered": "已掌握"}
        rows = []
        for c in cards:
            rows.append([
                c.subject, c.question, c.answer,
                mastery_map.get(c.mastery_level, c.mastery_level),
                str(c.review_count),
                ", ".join(c.tags or []),
                str(c.next_review_date) if c.next_review_date else "",
                str(c.created_at.date()) if c.created_at else ""
            ])
    return headers, rows


def _mistake_rows(mistakes, questions_only=False):
    diff_map = {"easy": "简单", "medium": "中等", "hard": "困难"}
    if questions_only:
        headers = ["科目", "题目", "难度", "错误次数", "标签"]
        rows = []
        for m in mistakes:
            rows.append([
                m.subject, m.question,
                diff_map.get(m.difficulty, m.difficulty),
                str(m.error_count),
                ", ".join(m.tags or []),
            ])
    else:
        headers = ["科目", "题目", "答案", "错误分析", "难度", "错误次数", "正确次数", "是否掌握", "标签", "下次复习日期", "创建日期"]
        rows = []
        for m in mistakes:
            rows.append([
                m.subject, m.question, m.answer, m.analysis or "",
                diff_map.get(m.difficulty, m.difficulty),
                str(m.error_count), str(m.correct_count),
                "已掌握" if m.mastered == "1" else "未掌握",
                ", ".join(m.tags or []),
                str(m.next_review_date) if m.next_review_date else "",
                str(m.created_at.date()) if m.created_at else ""
            ])
    return headers, rows


# ========== Cards Export ==========

@router.get("/cards/csv")
def export_cards_csv(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    mastery_level: str = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    cards = _query_cards(db, plan_id, user_id, subject, tag, mastery_level)
    headers, rows = _card_rows(cards, questions_only)
    csv_content = _generate_csv(headers, rows)
    csv_bytes = ("\ufeff" + csv_content).encode("utf-8")
    return StreamingResponse(
        BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=knowledge_cards.csv"}
    )


@router.get("/cards/excel")
def export_cards_excel(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    mastery_level: str = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    cards = _query_cards(db, plan_id, user_id, subject, tag, mastery_level)
    headers, rows = _card_rows(cards, questions_only)
    excel_bytes = _generate_excel(headers, rows, "知识卡片")
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=knowledge_cards.xlsx"}
    )


@router.get("/cards/pdf")
def export_cards_pdf(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    mastery_level: str = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    cards = _query_cards(db, plan_id, user_id, subject, tag, mastery_level)
    headers, rows = _card_rows(cards, questions_only)
    pdf_bytes = _generate_pdf(headers, rows, "知识卡片导出")
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=knowledge_cards.pdf"}
    )


# ========== Mistakes Export ==========

@router.get("/mistakes/csv")
def export_mistakes_csv(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    difficulty: str = Query(None),
    mastered: str = Query(None),
    min_errors: int = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    mistakes = _query_mistakes(db, plan_id, user_id, subject, tag, difficulty, mastered, min_errors)
    headers, rows = _mistake_rows(mistakes, questions_only)
    csv_content = _generate_csv(headers, rows)
    csv_bytes = ("\ufeff" + csv_content).encode("utf-8")
    return StreamingResponse(
        BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=mistakes.csv"}
    )


@router.get("/mistakes/excel")
def export_mistakes_excel(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    difficulty: str = Query(None),
    mastered: str = Query(None),
    min_errors: int = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    mistakes = _query_mistakes(db, plan_id, user_id, subject, tag, difficulty, mastered, min_errors)
    headers, rows = _mistake_rows(mistakes, questions_only)
    excel_bytes = _generate_excel(headers, rows, "错题本")
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mistakes.xlsx"}
    )


@router.get("/mistakes/pdf")
def export_mistakes_pdf(
    plan_id: UUID = Query(...),
    subject: str = Query(None),
    tag: str = Query(None),
    difficulty: str = Query(None),
    mastered: str = Query(None),
    min_errors: int = Query(None),
    questions_only: bool = Query(False),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    mistakes = _query_mistakes(db, plan_id, user_id, subject, tag, difficulty, mastered, min_errors)
    headers, rows = _mistake_rows(mistakes, questions_only)
    pdf_bytes = _generate_pdf(headers, rows, "错题本导出")
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=mistakes.pdf"}
    )


# ========== Tags & Subjects Query (for export filter cascade) ==========

@router.get("/tags")
def get_tags_by_subject(
    plan_id: UUID = Query(...),
    type: str = Query("cards"),
    subject: str = Query(None),
    user_id: UUID = Depends(_get_user_id),
    db: Session = Depends(get_db)
):
    """Get available subjects and tags (cascaded by subject) for export filters."""
    plan = db.query(StudyPlan).filter(StudyPlan.id == plan_id, StudyPlan.user_id == user_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")

    result = {"subjects": [], "tags_by_subject": {}}

    if type == "cards":
        items = db.query(FlashCard).filter(FlashCard.plan_id == plan_id).all()
        for item in items:
            subj = item.subject
            tags = item.tags or []
            if subj not in result["subjects"]:
                result["subjects"].append(subj)
                result["tags_by_subject"][subj] = set()
            for t in tags:
                result["tags_by_subject"][subj].add(t)
    else:
        items = db.query(Mistake).filter(Mistake.plan_id == plan_id).all()
        for item in items:
            subj = item.subject
            tags = item.tags or []
            if subj not in result["subjects"]:
                result["subjects"].append(subj)
                result["tags_by_subject"][subj] = set()
            for t in tags:
                result["tags_by_subject"][subj].add(t)

    if subject and subject in result["tags_by_subject"]:
        result["tags"] = sorted(list(result["tags_by_subject"].get(subject, set())))
    else:
        all_tags = set()
        for s in result["tags_by_subject"]:
            all_tags.update(result["tags_by_subject"][s])
        result["tags"] = sorted(list(all_tags))

    # Convert sets to lists for JSON
    result["tags_by_subject"] = {k: sorted(list(v)) for k, v in result["tags_by_subject"].items()}
    return result
